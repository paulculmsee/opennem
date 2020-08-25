import json
from datetime import datetime
from itertools import groupby
from pprint import pprint
from typing import List, Optional, Union

from openpyxl import load_workbook

from opennem.core.dispatch_type import parse_dispatch_type
from opennem.core.facilitystatus import map_aemo_facility_status
from opennem.core.fueltechs import lookup_fueltech
from opennem.core.loader import PROJECT_DATA_PATH
from opennem.core.normalizers import (
    clean_capacity,
    normalize_aemo_region,
    normalize_duid,
    station_name_cleaner,
)
from opennem.core.stations.station_code_from_duids import (
    station_code_from_duids,
)
from opennem.core.stations.station_name_code_map import station_name_code_map
from opennem.core.unit_codes import get_basecode, get_unit_code
from opennem.core.unit_parser import parse_unit_duid
from opennem.db.load_fixtures import load_fixture
from opennem.exporter.encoders import OpenNEMJSONEncoder

from .mms import run_import_mms

participant_keys = ["name", "abn"]

FACILITY_KEYS = [
    "Region",
    "Status",
    "station_name",
    "owner",
    "TechType",
    "FuelType",
    "duid",
    "unit_num",
    "unit_capacity_lower",
    "unit_capacity",
    "capacity_upper",
    "capacity_registered",
    "capacity_storage",
    "status",
    "dispatch_type",
    "UseDate",
    "ClosureDateExpected",
    "ClosureDate",
    "SurveyID",
    "SurveyVersion",
    "SurveyEffective",
]


def parse_comissioned_date(
    date_str: Union[str, datetime]
) -> Optional[datetime]:
    dt = None

    if type(date_str) is datetime:
        dt = date_str

    try:
        if type(date_str) is str:
            dt = datetime.strptime(date_str, "%d/%m/%y")
    except ValueError:
        logger.error("Error parsing date: {}".format(facility_comissioned))

    return dt


def lookup_station_code(
    duids: List[str], station_name: str, station_code_map: dict
) -> Optional[str]:

    station_code = None
    station_name = station_name.strip()

    for facility_duid in duids:
        if facility_duid in station_code_map:
            station_code = station_code_map[facility_duid]
            continue

    if not station_code:
        station_code = station_name_code_map(station_name)

    if not station_code:
        station_code = station_code_from_duids(duids)

    return station_code


FACILITY_INVALID_STATUS = [
    "Publically Announced",
    "Upgrade",
    "Emerging",
    "Expansion",
    "Maturing",
]


def gi_filter(record):
    if record["station_name"] == None:
        return False

    if record["FuelType"] in ["Natural Gas Pipeline"]:
        return False

    # skip these statuses too
    if record["status"] in FACILITY_INVALID_STATUS:
        return False

    return True


capacity_keys = [
    "unit_capacity_lower",
    "unit_capacity",
    "capacity_upper",
    "capacity_registered",
    "capacity_storage",
]


def get_capacities(record) -> dict:
    _d = {}

    for capacity_key in capacity_keys:
        _d[capacity_key] = clean_capacity(record[capacity_key])

    return _d


def gi_grouper(records, station_code_map):

    # filter out records we don't want
    records = list(filter(gi_filter, records))

    records = [
        {"name": station_name_cleaner(i["station_name"]), **i} for i in records
    ]

    grouped_records = {}

    for _, v in groupby(
        records, key=lambda v: (v["name"].strip(), v["owner"].strip())
    ):
        v = list(v)

        key = v[0]["station_name"]

        if not key in grouped_records:
            grouped_records[key] = []

        grouped_records[key] += v

    records_parsed = []

    for station_name, facilities in grouped_records.items():
        facility_index = 0

        for i in facilities:
            name = station_name_cleaner(i["station_name"])
            duid = normalize_duid(i["duid"])

            unit = parse_unit_duid(i["unit_num"], duid)
            units_num = i["unit_num"] or 1
            unit_id = facility_index + (units_num - 1)
            unit = parse_unit_duid(unit_id, duid)

            fueltech = lookup_fueltech(i["FuelType"], techtype=i["TechType"])

            if not duid:
                duid = get_unit_code(unit, duid, i["station_name"])

            facility_duids = [duid]
            station_code = lookup_station_code(
                facility_duids, station_name, station_code_map
            )

            records_parsed.append(
                {
                    # not a real station id
                    "id": i["SurveyID"],
                    "name": name,
                    "network_code": duid,
                    "network_name": i["station_name"],
                    "status": "operating",
                    "station_code": station_code,
                    "network_region": normalize_aemo_region(i["Region"]),
                    "network_name": i["station_name"].strip(),
                    "fueltech": fueltech,
                    "status": map_aemo_facility_status(i["status"]),
                    "registered": parse_comissioned_date(i["SurveyEffective"]),
                    **get_capacities(i),
                }
            )

            facility_index += 1

    coded_records = {}

    for station_code, facilities in groupby(
        records_parsed, key=lambda x: x["station_code"]
    ):

        # station_name = facilities[0]["name"]
        facilities = list(facilities)

        if not station_code:
            raise Exception(
                "Unmapped station {}: {}".format(station_code, facilities)
            )

        if not station_code in coded_records:
            coded_records[station_code] = {
                "station_code": station_code,
                # "name": station_name_cleaner(station_name),
                # "network_name": station_name,
                "facilities": [],
            }

        coded_records[station_code]["facilities"] += facilities

    return coded_records


def load_gi():
    aemo_path = PROJECT_DATA_PATH / "aemo" / "nem_gi_202007.xlsx"

    if not aemo_path.is_file():
        raise Exception("Not found: {}".format(aemo_path))

    wb = load_workbook(aemo_path, data_only=True)

    generator_ws = wb["ExistingGeneration&NewDevs"]

    records = []

    for row in generator_ws.iter_rows(min_row=3, values_only=True):
        row_collapsed = row[0:10] + (row[11],) + row[12:19] + row[21:]

        return_dict = dict(zip(FACILITY_KEYS, list(row_collapsed)))

        records.append(return_dict)

    return records


def run_import_nem_gi():
    station_fixture = load_fixture("facility_registry.json")
    mms = run_import_mms()

    nem_stations_registry = {
        i: v
        for i, v in station_fixture.items()
        if v["location"]["state"] not in ["WA"]
    }

    nem_gi = load_gi()

    mms_duid_station_map = {}

    for station, station_record in mms.items():
        for network_code in [
            i["network_code"] for i in station_record["facilities"]
        ]:
            mms_duid_station_map[network_code] = station

    nem_gi = gi_grouper(nem_gi, mms_duid_station_map)

    with open("data/nem_gi.json", "w") as fh:
        json.dump(nem_gi, fh, indent=4, cls=OpenNEMJSONEncoder)


if __name__ == "__main__":
    run_import_nem_gi()
